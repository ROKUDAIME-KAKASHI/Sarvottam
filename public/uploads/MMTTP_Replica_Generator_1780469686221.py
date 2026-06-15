
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def generate_mmttp_calendar(data, month_year, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Online"

    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 40

    ws.merge_cells("B2:E5")
    title = f"Malaviya Mission Teacher Training Programme\n{month_year} Calendar\nhttps://mmc.ugc.ac.in/Home/DefaultPage"
    ws["B2"] = title
    ws["B2"].font = Font(size=14, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = ["S.No.", "Category", "Details", "Organising University"]
    for col, text in zip(["B", "C", "D", "E"], headers):
        ws[f"{col}6"] = text
        ws[f"{col}6"].font = Font(bold=True)
        ws[f"{col}6"].border = thin
        ws[f"{col}6"].alignment = Alignment(horizontal="center")

    row = 7
    sno = 1

    for category, programs in data.items():
        start_row = row

        for p in programs:
            block_start = row

            ws.cell(row=row, column=4, value=f"Category: {category}")
            ws.cell(row=row, column=5, value=p["university"])
            row += 1

            ws.cell(row=row, column=4, value=f"Title: {p['title']}")
            row += 1

            ws.cell(row=row, column=4, value=f"Start: {p['start']}")
            row += 1

            ws.cell(row=row, column=4, value=f"End: {p['end']}")
            row += 1

            ws.cell(row=row, column=4, value=f"Mode: {p['mode']}")
            row += 1

            ws.merge_cells(start_row=block_start, start_column=5,
                           end_row=row, end_column=5)

            row += 1

        end_row = row - 1

        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=end_row, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3,
                       end_row=end_row, end_column=3)

        ws.cell(start_row, 2, sno)
        ws.cell(start_row, 3, category)

        for r in range(start_row, end_row + 1):
            for c in range(2, 6):
                ws.cell(r, c).border = thin

        sno += 1

    wb.save(output_file)


if __name__ == "__main__":
    sample_data = {
        "Guru Dakshta(Faculty Induction Programme)": [
            {
                "title": "Faculty Induction Programme",
                "start": "21-Apr-2026",
                "end": "19-May-2026",
                "mode": "Online",
                "university": "Bangalore University"
            }
        ],
        "Refresher Course": [
            {
                "title": "AI Literacy",
                "start": "13-Apr-2026",
                "end": "25-Apr-2026",
                "mode": "Online",
                "university": "Rani Durgawati Vishwavidyalaya"
            }
        ]
    }

    generate_mmttp_calendar(
        sample_data,
        "April-26",
        "MMTTP_Calendar.xlsx"
    )
